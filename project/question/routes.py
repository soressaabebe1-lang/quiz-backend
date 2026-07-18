from io import BytesIO

from flask import Blueprint, jsonify, request
from flask_jwt_extended import get_jwt_identity, jwt_required
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from project.models import Questions, db
from project.utils import admin_required

question_bp = Blueprint("question", __name__)


def _normalize_header(value):
    return "".join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def parse_questions_from_excel(file_storage):
    if not file_storage:
        raise ValueError("No file uploaded")

    filename = (file_storage.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported")

    workbook = load_workbook(BytesIO(file_storage.read()), data_only=True)
    if not workbook.sheetnames:
        raise ValueError("Workbook is empty")

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty")

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    header_positions = {}
    for index, header in enumerate(headers):
        normalized = _normalize_header(header)
        if normalized in {"question", "questiontext", "q"}:
            header_positions["question"] = index
        elif normalized in {"a", "optiona", "choicea"}:
            header_positions["A"] = index
        elif normalized in {"b", "optionb", "choiceb"}:
            header_positions["B"] = index
        elif normalized in {"c", "optionc", "choicec"}:
            header_positions["C"] = index
        elif normalized in {"d", "optiond", "choiced"}:
            header_positions["D"] = index
        elif normalized in {"answer", "correctanswer", "correct"}:
            header_positions["answer"] = index

    missing = {"question", "A", "B", "C", "D", "answer"} - set(header_positions)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    parsed_questions = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue

        item = {}
        for field, index in header_positions.items():
            value = row[index] if index < len(row) else None
            item[field] = "" if value is None else str(value).strip()

        if not item["question"] or not item["A"] or not item["B"] or not item["C"] or not item["D"] or not item["answer"]:
            continue

        item["answer"] = item["answer"].strip().upper()
        parsed_questions.append(item)

    return parsed_questions


@question_bp.route("/question", methods=["GET"])
@jwt_required()
def get_questions():
    question_items = Questions.query.all()
    questions = [items.to_dict() for items in question_items]
    return jsonify({"data": questions}), 200


@question_bp.route("/question", methods=["POST"])
@admin_required
def add_question():
    data = request.get_json()
    if not data or not data.get("question") or not data.get("A") or not data.get("B") or not data.get("C") or not data.get("D") or not data.get("answer"):
        return jsonify({"message": "Incomplete data"}), 400

    existing_question = Questions.query.filter_by(question=data["question"]).first()

    if existing_question:
        return jsonify({"msg": "question already exists"}), 400

    new_questions = Questions(
        question=data["question"],
        A=data["A"],
        B=data["B"],
        C=data["C"],
        D=data["D"],
        answer=data["answer"],
    )

    db.session.add(new_questions)
    db.session.commit()
    user_id = get_jwt_identity()
    return jsonify({"status": "created", "created by": user_id, "message": "question is added"}), 201


@question_bp.route("/question/import", methods=["POST"])
@admin_required
def import_questions():
    file = request.files.get("file")
    if not file:
        return jsonify({"message": "No file uploaded"}), 400

    try:
        parsed_questions = parse_questions_from_excel(file)
    except ValueError as error:
        return jsonify({"message": str(error)}), 400
    except InvalidFileException:
        return jsonify({"message": "Invalid Excel file"}), 400

    created = 0
    skipped = 0
    for item in parsed_questions:
        existing_question = Questions.query.filter(Questions.question.ilike(item["question"])).first()
        if existing_question:
            skipped += 1
            continue

        question = Questions(
            question=item["question"],
            A=item["A"],
            B=item["B"],
            C=item["C"],
            D=item["D"],
            answer=item["answer"],
        )
        db.session.add(question)
        created += 1

    db.session.commit()
    return jsonify({"status": "imported", "created": created, "skipped": skipped, "message": "questions imported"}), 201


@question_bp.route("/question/<int:id>", methods=["DELETE"])
@admin_required
def delete_question(id):
    question = Questions.query.get(id)
    if not question:
        return jsonify({"message": "Question not found"}), 404
    db.session.delete(question)
    db.session.commit()
    return jsonify({"status": "deleted"}), 200


@question_bp.route("/question/<int:id>", methods=["GET"])
@jwt_required()
def get_question(id):
    question = Questions.query.get(id)
    if not question:
        return jsonify({"message": "Not found"}), 404
    return jsonify({"data": question.to_dict()}), 200